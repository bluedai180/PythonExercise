from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment, NamedStyle
import os
import tkinter.filedialog as filedialog
from tkinter import *

"""
作者：Blue（应用开发部）
功能：统计MonkeyLog的crash和anr信息
版本：1.0
"""


class MonkeyLog(object):

    log_path = ''
    info_total = []
    info_final = []

    def __init__(self):
        if os.path.exists(os.path.dirname(os.path.abspath(self.log_path)) + '\log.xlsx'):
            os.remove(os.path.dirname(os.path.abspath(self.log_path)) + '\log.xlsx')

    def get_info(self, app_name, message, index):
        name = []
        detail = []
        info_index = []
        info_message = []
        info_total_message = []
        f = open(self.log_path, 'r', encoding='utf-8')
        for line, data in enumerate(f):
            if data.startswith(app_name):
                name.append(data.split(' ')[2].replace('\n', ''))
                info_index = range(line + 6, line + 50)
                if len(info_message) != 0:
                    info_total_message.append("".join(info_message))
                    info_message.clear()
                continue
            if data.startswith(message):
                detail.append(' '.join(data.split(' ')[index:]).replace('\n', ''))
                continue
            if line in info_index:
                if data.startswith('//'):
                    info_message.append(data)
        info_total_message.append("".join(info_message))

        if len(name) != len(detail):
            print("Wrong message !!!!!!!")
        if len(name) != len(info_total_message):
            print("Wrong message")

        for i in range(len(name)):
            temp = [name[i], detail[i], info_total_message[i]]
            self.info_total.append(temp)
        f.close()

    def analyze_info(self):
        self.info_total.sort(key=lambda x: x[0])
        each_bug_num = self.count_bug(self.info_total)
        each_app_num = self.count_app(self.info_total)
        for i in range(len(each_app_num)):
            self.info_total[i].insert(0, each_app_num[i])
        for i in range(len(each_bug_num)):
            self.info_total[i].append(each_bug_num[i])
        for x in self.info_total:
            if x not in self.info_final:
                self.info_final.append(x)


    def write_excel(self, ws, title_style, content_style, content_long_style):
        ws.append(['总数', '进程名', '错误信息', '详细信息', '复现次数'])
        merge_line_num = self.count_merge(self.info_final)
        merge_line_bug_num = self.count_merge_bug(self.info_final)
        print(merge_line_bug_num)
        print(merge_line_num)
        length = len(merge_line_num)
        for i in range(length):
            if length == 1:
                ws.merge_cells('A%s:A%s' % (2, merge_line_num[-1]))
                ws.merge_cells('B%s:B%s' % (2, merge_line_num[-1]))
            elif i < length - 1:
                if merge_line_num[i + 1] - merge_line_num[i] != 1:
                    ws.merge_cells('A%s:A%s' % (merge_line_num[i] + 1, merge_line_num[i + 1]))
                    ws.merge_cells('B%s:B%s' % (merge_line_num[i] + 1, merge_line_num[i + 1]))
                if merge_line_bug_num[i + 1] - merge_line_bug_num[i] != 1:
                    ws.merge_cells('C%s:C%s' % (merge_line_bug_num[i] + 1, merge_line_bug_num[i + 1]))
        for x in self.info_final:
            ws.append(x)
        self.format_excel(ws, title_style, content_style, content_long_style)
        wb.save(os.path.dirname(os.path.abspath(self.log_path)) + "\log.xlsx")
        self.info_final.clear()
        self.info_total.clear()
        print("%s表数据已经写入." % ws.title)

    def count_app(self, info):
        temp_name = []
        temp_num = []
        for x in info:
            temp_name.append(x[0])
        for x in temp_name:
            temp_num.append(temp_name.count(x))
        return temp_num

    def count_bug(self, info):
        temp_num = []
        for x in info:
            temp_num.append(self.info_total.count(x))
        return temp_num

    def count_merge(self, info):
        temp_num = []
        length = len(info)
        for i in range(length):
            if i < length - 1:
                if info[i][1] != info[i + 1][1]:
                    temp_num.append(i + 2)
            else:
                temp_num.append(length + 1)
        return temp_num

    def count_merge_bug(self, info):
        temp_num = []
        length = len(info)
        for i in range(length):
            if i <length - 1:
                if info[i][2] != info[i+1][2]:
                    temp_num.append(i + 2)
            else:
                temp_num.append(length + 1)
        return temp_num

    def format_excel(self, ws, title_style, content_style, content_long_style):
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 100
        ws.column_dimensions['E'].width = 10
        for i in range(ws.max_row):
            ws.row_dimensions[i + 1].height = 60
        for x in ws[1]:
            x.style = title_style
        for x in ws['A:B']:
            for y in x:
                y.style = content_style
        for x in ws['C'][1:]:
            x.style = content_long_style
        for x in ws['D'][1:]:
            x.style = content_long_style
        for x in ws['E'][1:]:
            x.style = content_style

    @staticmethod
    def open_win():
        root = Tk()
        root.title("MonkeyLog 分析")
        ws = root.winfo_screenwidth()
        hs = root.winfo_screenheight()
        x = ws/2 - 400/2
        y = hs/2 - 200/2
        root.geometry("400x200+%d+%d" % (x, y))

        def callback():
            MonkeyLog.log_path = filedialog.askopenfilename()
            entry.insert(0, MonkeyLog.log_path)
        button = Button(root, text="选择MonkeyLog文件", command=callback)
        quit_btn = Button(root, text="确定", command=root.destroy)
        entry = Entry(root)
        entry.pack(side=TOP, anchor="nw", fill=X, pady=40)
        button.pack(side=TOP)
        quit_btn.pack(side=TOP, pady=10)
        root.mainloop()

if __name__ == '__main__':
    MonkeyLog.open_win()

    crash = MonkeyLog()
    anr = MonkeyLog()
    wb = Workbook()

    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    title = NamedStyle(name="title")
    title.font = Font(name=u'宋体', size=11)
    title.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title.border = Border(left=left, right=right, top=top, bottom=bottom)
    content = NamedStyle(name="content")
    content.font = Font(name=u'宋体', size=11)
    content.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    content.border = Border(left=left, right=right, top=top, bottom=bottom)
    content_long = NamedStyle(name="content_long")
    content_long.font = Font(name=u'宋体', size=11)
    content_long.border = Border(left=left, right=right, top=top, bottom=bottom)
    content_long.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    crash.get_info('// CRASH: ', '// Long Msg:', 3)
    crash.analyze_info()
    ws_crash = wb.active
    ws_crash.title = "crash"
    crash.write_excel(ws_crash, title, content, content_long)

    # anr.get_info('ANR in', 'Reason:', 1)
    # anr.analyze_info()
    # ws_anr = wb.create_sheet("anr")
    # anr.write_excel(ws_anr, title, content, content_long)
    print("数据已经导出完毕，请到其log文件目录下查看log.xlsx文件！")



