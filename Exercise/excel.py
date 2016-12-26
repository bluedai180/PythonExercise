import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font, Alignment, NamedStyle
import datetime
import sys

'''
文档：https://openpyxl.readthedocs.io/en/default/
功能：整理组内日报脚本
作者：bluedai180
时间：2016/12/23
'''


class Daily(object):
    name = []
    day_index = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
    file_path = '../Files/excel_files'
    file_name = '../Files/excel_files/应用一课日报.xlsx'
    excel_title = '应用一课日报'
    excel_file = []
    excel_name = []

    def __init__(self):
        self.wb_new = Workbook()
        self.ws_new = self.wb_new.active
        self.ws_new.title = self.excel_title
        if (os.path.exists(self.file_name)):
            os.remove(self.file_name)

    def check_file(self):
        self.excel_file = os.listdir(self.file_path)
        self.excel_name = []
        name = []
        for x in self.excel_file:
            self.excel_name.append(x.split('_')[1][0:-5])
        for x in self.name:
            if x not in self.excel_name:
                name.append(x)
        if len(name) is not 0:
            print("以下人员未交日报：\n")
            print(", ".join(name))
        else:
            print("全员日报已交齐！")

    def do_files(self, commend):
        self.ws_new.append(['项目', '工作类别', 'Bug ID', '简要描述', '优先级', '是否reopen', 'reopen原因', '解决方案',
                            '原因', '责任人', '日期', '备注'])
        day = self.day_index[commend - 1]
        for x in self.excel_file:
            print("正在读取: %s" % x)
            wb = load_workbook('%s/%s' % (self.file_path, x), read_only=True)
            ws = wb.get_sheet_by_name(day)
            excel_cell = list(ws.rows)
            temp = []
            for x in excel_cell[1:]:
                if x[0].value is None:
                    break
                for y in x:
                    if type(y.value) is datetime.datetime:
                        temp.append(y.value.strftime("%Y/%m/%d"))
                    else:
                        temp.append(y.value)
                self.ws_new.append(temp)
                temp.clear()

        self.format_file()
        self.check_name()
        self.wb_new.save(self.file_name)


    def format_file(self):
        self.ws_new.column_dimensions['A'].width = 20
        self.ws_new.column_dimensions['B'].width = 14
        self.ws_new.column_dimensions['C'].width = 14
        self.ws_new.column_dimensions['D'].width = 70
        self.ws_new.column_dimensions['E'].width = 10
        self.ws_new.column_dimensions['F'].width = 14
        self.ws_new.column_dimensions['G'].width = 14
        self.ws_new.column_dimensions['H'].width = 14
        self.ws_new.column_dimensions['I'].width = 14
        self.ws_new.column_dimensions['J'].width = 14
        self.ws_new.column_dimensions['K'].width = 14
        self.ws_new.column_dimensions['L'].width = 25

        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        title = NamedStyle(name="title")
        title.font = Font(name=u'宋体', size=11, bold=True)
        title.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title.border = Border(left=left, right=right, top=top, bottom=bottom)
        content = NamedStyle(name="content")
        content.font = Font(name=u'宋体', size=11)
        content.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        content.border = Border(left=left, right=right, top=top, bottom=bottom)
        content_long = NamedStyle(name="content_long")
        content_long.font = Font(name=u'宋体', size=11)
        content_long.border = Border(left=left, right=right, top=top, bottom=bottom)
        content_long.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for i in range(self.ws_new.max_row):
            self.ws_new.row_dimensions[i + 1].height = 40

        for x in self.ws_new['A:C']:
            for y in x:
                y.style = content

        for x in self.ws_new['E:F']:
            for y in x:
                y.style = content

        for x in self.ws_new['H:K']:
            for y in x:
                y.style = content

        for x in self.ws_new[1]:
            x.style = title

        for x in self.ws_new['D'][1:]:
            x.style = content_long

        for x in self.ws_new['L'][1:]:
            x.style = content_long

        for x in self.ws_new['G'][1:]:
            x.style = content_long

    def check_name(self):
        content_name = []
        name = []
        for x in self.ws_new['J'][1:]:
            content_name.append(x.value)
        for x in self.excel_name:
            if x not in content_name:
                name.append(x)
        if len(name) is not 0:
            print("\n以下人员已提交日报，但工作表内容为空:\n")
            print(", ".join(name))

if __name__ == '__main__':
    daily = Daily()
    daily.check_file()
    is_do = input("\n是否继续生成日报？[y/n]:")
    if is_do == 'y':
        commend = input("\n需要输出星期几的日报: \n1. 星期一\n2. 星期二\n3. 星期三\n4. 星期四\n5. 星期五\n6. 星期六\n"
                        "7. 星期日\n请输入序号： ")
        daily.do_files(int(commend))
        print("\n=======请查看生成的日报汇总。=======")
    elif is_do == 'n':
        sys.exit(0)
    else:
        print("请输入y或者n！")



















