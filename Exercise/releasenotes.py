import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment, NamedStyle
import tkinter.filedialog as filedialog
from tkinter import *

"""
===================================================
作者：blue
部门：北京基地应用开发部
===================================================
版本：v1.0
功能：对svn的ReleaseNotes进行分析处理，输出成excel文件
===================================================
版本：v1.1
功能：
1.新增统计正确与错误的个数
2.对于格式错误的提交不会影响正确格式提交的excel表数据
3.新增“格式错误列表”，统计格式错误的信息，并附修改建议
===================================================
"""


class ReleaseNotes(object):
    path = ""

    """初始化方法，创建处理Excel的对象"""
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "svn_info"
        self.ws.append(['Bug ID', '责任人', '问题描述', '原因分析', '修改方案', 'svn 版本'])
        self.ws_wrong = self.wb.create_sheet("格式错误的提交")
        self.ws_wrong.append(['错误类型', 'svn 版本', '提交人', '详细信息', '修改意见'])

    """对txt文件进行分隔，获取每一条信息"""
    def get_each_info(self):
        lines = []
        info = []
        file = open(self.path, 'r', encoding='utf-8')
        for data in file:
            lines.append(data)
            if data.startswith("-------"):
                info.append(
                    "".join(lines).replace("------------------------------------------------------------------------",
                                           ""))
                lines.clear()
        info_total = [x for x in info if x != '\n']
        file.close()
        print("处理结果：")
        self.get_detail_info(info_total)

    """对每一条信息进行内容提取，并填充到ws"""
    def get_detail_info(self, info_total):
        temp = []
        temp_wrong = []
        for text in info_total:
            text_list = [x for x in text.split('\n') if x != ""]
            try:
                author = text_list[1].split(']')[0].split(":")[1].lstrip()
                bug_id = text_list[1].split(']')[1].split("+")[1].lstrip()
                svn_id = text_list[0].split('|')[0].replace('r', '').lstrip()
                temp.append(bug_id)
                temp.append(author)
                a = text_list.index("1.问题描述")
                b = text_list.index("2.原因分析")
                if "问题描述" in text:
                    temp.append("".join(text_list[a + 1:b]).lstrip())
                c = text_list.index("3.修改方案")
                if "原因分析" in text:
                    temp.append("".join(text_list[b + 1:c]).lstrip())
                d = text_list.index("4.影响范围")
                if "修改方案" in text:
                    temp.append("".join(text_list[c + 1:d]).lstrip())
                temp.append(svn_id)
            except ValueError:
                svn_id = text_list[0].split('|')[0].replace('r', '').lstrip()
                author_name = text_list[0].split('|')[1].lstrip()
                temp_wrong.append("小标题格式错误")
                temp_wrong.append(svn_id)
                temp_wrong.append(author_name)
                temp_wrong.append("\n".join(text_list[1:]))
                temp_wrong.append("标题文字要正确、删除多余的空格、删除冒号、注意是“修改方案”不是“解决方案”")
                self.ws_wrong.append(temp_wrong)
                temp_wrong.clear()
            except IndexError:
                svn_id = text_list[0].split('|')[0].replace('r', '').lstrip()
                author_name = text_list[0].split('|')[1].lstrip()
                temp_wrong.append("总标题格式错误")
                temp_wrong.append(svn_id)
                temp_wrong.append(author_name)
                temp_wrong.append("\n".join(text_list[1:]))
                temp_wrong.append("标准格式：[Author : daiqingchen] [YL50B71_CMCC + ODMAAD-394]、中括号要全、不要省略+号")
                self.ws_wrong.append(temp_wrong)
                temp_wrong.clear()
            finally:
                if len(temp) == 6:
                    self.ws.append(temp)
                temp.clear()

    """将数据写入Excel，并设置单元格样式"""
    def write_excel(self):
        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        title = NamedStyle(name="title")
        title.font = Font(name=u'宋体', size=11, bold=True)
        title.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title.border = Border(left=left, right=right, top=top, bottom=bottom)
        content = NamedStyle(name="content")
        content.font = Font(name=u'宋体', size=11)
        content.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        content.border = Border(left=left, right=right, top=top, bottom=bottom)
        content_long = NamedStyle(name="content_long")
        content_long.font = Font(name=u'宋体', size=11)
        content_long.border = Border(left=left, right=right, top=top, bottom=bottom)
        content_long.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.ws.column_dimensions['A'].width = 15
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 80
        self.ws.column_dimensions['D'].width = 55
        self.ws.column_dimensions['E'].width = 55
        self.ws.column_dimensions['F'].width = 10
        self.ws_wrong.column_dimensions['A'].width = 20
        self.ws_wrong.column_dimensions['B'].width = 15
        self.ws_wrong.column_dimensions['C'].width = 15
        self.ws_wrong.column_dimensions['D'].width = 80
        self.ws_wrong.column_dimensions['E'].width = 80
        for i in range(self.ws.max_row):
            self.ws.row_dimensions[i + 1].height = 30
        for i in range(self.ws_wrong.max_row):
            self.ws_wrong.row_dimensions[i + 1].height = 50
        for x in self.ws[1]:
            x.style = title
        for x in self.ws_wrong[1]:
            x.style = title
        for x in self.ws['A'][1:]:
            x.style = content
        for x in self.ws_wrong['A'][1:]:
            x.style = content
        for x in self.ws['B'][1:]:
            x.style = content
        for x in self.ws_wrong['B'][1:]:
            x.style = content
        for x in self.ws['C'][1:]:
            x.style = content_long
        for x in self.ws_wrong['C'][1:]:
            x.style = content
        for x in self.ws['D'][1:]:
            x.style = content_long
        for x in self.ws_wrong['D'][1:]:
            x.style = content_long
        for x in self.ws['E'][1:]:
            x.style = content_long
        for x in self.ws_wrong['E'][1:]:
            x.style = content_long
        for x in self.ws['F'][1:]:
            x.style = content

        print("格式正确个数 %d" % self.ws.max_row)
        print("格式错误个数 %d" % self.ws_wrong.max_row)
        self.wb.save(os.path.dirname(os.path.abspath(self.path)) + "\\" + os.path.basename(self.path).replace(".txt",
                                                                                                              ".xlsx"))

    """静态方法，打开win窗口"""
    @staticmethod
    def open_win():
        root = Tk()
        root.title("svn release notes 分析")
        ws = root.winfo_screenwidth()
        hs = root.winfo_screenheight()
        x = ws / 2 - 400 / 2
        y = hs / 2 - 200 / 2
        root.geometry("400x200+%d+%d" % (x, y))

        def callback():
            ReleaseNotes.path = filedialog.askopenfilename()
            entry.insert(0, ReleaseNotes.path)

        button = Button(root, text="选择ReleaseNotes文件", command=callback)
        quit_btn = Button(root, text="确定", command=root.destroy)
        entry = Entry(root)
        entry.pack(side=TOP, anchor="nw", fill=X, pady=40)
        button.pack(side=TOP)
        quit_btn.pack(side=TOP, pady=10)
        root.mainloop()

if __name__ == "__main__":
    ReleaseNotes.open_win()
    notes = ReleaseNotes()
    notes.get_each_info()
    notes.write_excel()
    print("数据已经导出完毕，请到对应目录下查看excel文件！")
